import io
import csv
import json
import datetime

from openpyxl import load_workbook
from django.shortcuts import render
from django.shortcuts import render_to_response
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.http import HttpResponseRedirect

from controlpanel.utils.csv_parser import CSVParser



def handle_404(request):
  return HttpResponseRedirect("/")

'''
Shell to serve angular app,
all logic is then worked on the API
'''


def web_app(request):
    return render_to_response("index.html")


'''
Control Panel Login Page
'''


def app_login(request):
    return render_to_response("controlpanel/login.html")


'''
Parse Excel Mass Event Invite
'''

@csrf_exempt
def import_excel_event(request):
    csv_file = request.FILES['file']
    decoded_file = csv_file.read().decode('utf-8')
    io_string = io.StringIO(decoded_file)

    parser = CSVParser(io_string)
    parsed_data = parser.all_fields_required()

    for item in parsed_data:
        date_columns = ['start_date', 'end_date']

        for column in date_columns:
            if parser.validate_date_format(item[column]) is not None:
                return JsonResponse({"error": '%s not in valid format' % column},
                                    safe=False, status=400)


        if parser.date_not_in_past(item['start_date']) is False:
            return JsonResponse({"error": 'Start date can not be in the past'},
                                safe=False, status=400)

        if parser.future_dates_is_greater_than_past(item['end_date'],
                                                    item['start_date']) is False:
            return JsonResponse({"error": 'Start date can not be greater than end date'},
                                safe=False, status=400)


    return JsonResponse(json.dumps(parsed_data), safe=False)
    # input_excel = request.FILES['file']

    # wb = load_workbook(filename=input_excel)

    # ws = wb.get_active_sheet()

    # event_dict = []

    # for row in ws.rows:
    #     event_info = []
    #     for col in row:
    #         if col.value is not None:
    #             event_info.append(col.value)
    #     if len(event_info):
    #         event_dict.append(event_info)

    # events = event_dict[1:]
    # column_titles = event_dict[:1]

    # column_titles = [title.lower() for title in column_titles[0]]

    # event_dict = []

    # for i in events:
    #     # all fields are required
    #     if len(i) is not len(column_titles):
    #         return JsonResponse({"error": "All fields are required"},
    #                             safe=False, status=500)

    #     # end date is not greater than start date
    #     if i[3] < i[2]:
    #         return JsonResponse({"error": "Start date can not be greater than end date"},
    #                             safe=False, status=500)

    #     # start date is not in the past
    #     if i[2] < datetime.datetime.now():
    #         return JsonResponse({"error": "Start date can not be in the past"},
    #                             safe=False, status=500)

    #     # check date time format (2017-05-12 09:00:00)
    #     try:
    #         datetime.datetime.strptime(str(i[2]), "%Y-%m-%d %H:%M:%S")
    #         datetime.datetime.strptime(str(i[3]), "%Y-%m-%d %H:%M:%S")
    #     except ValueError:
    #         return JsonResponse({"error": "Invalid date format"},
    #                             safe=False, status=500)

    #     i[2] = str(i[2])
    #     i[3] = str(i[3])

    #     event_dict.append(dict(zip(column_titles, i)))

    # return JsonResponse(json.dumps(event_dict), safe=False)

'''
Parse Excel Mass Announcements Import
'''
@csrf_exempt
def import_list(request):
    csv_file = request.FILES['file']
    decoded_file = csv_file.read().decode('utf-8')
    io_string = io.StringIO(decoded_file)

    parser = CSVParser(io_string)
    parsed_data = parser.all_fields_required()

    return JsonResponse(parsed_data, safe=False)

'''
Parse Clubs Mass Upload
'''
@csrf_exempt
def import_clubs(request):
    csv_file = request.FILES['file']
    decoded_file = csv_file.read().decode('utf-8')
    io_string = io.StringIO(decoded_file)

    parser = CSVParser(io_string)
    parsed_data = parser.all_fields_required()
    return JsonResponse(parsed_data, safe=False)

'''
Parse Services Excel Mass Upload
'''
@csrf_exempt
def import_excel_service(request):
    input_excel = request.FILES['file']

    wb = load_workbook(filename=input_excel)

    ws = wb.get_active_sheet()

    service_dict = []

    for row in ws.rows:
        service_info = []
        for col in row:
            if col.value is not None:
                service_info.append(col.value)
        if len(service_info):
            service_dict.append(service_info)

    services = service_dict[1:]
    column_titles = service_dict[:1]

    column_titles = [title.lower().replace(" ", "_") for title in column_titles[0]]

    service_dict = []

    for i in services:
        # all fields are required
        if len(i) is not len(column_titles):
            return JsonResponse({ "error": "All fields are required" }, safe=False, status=500)

        service_dict.append(dict(zip(column_titles, i)))

    return JsonResponse(json.dumps(service_dict), safe=False)
